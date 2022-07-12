import streamlit as st
import pandas as pd
import numpy as np
from pandas import ExcelWriter
import re
from snowflake.snowpark.session import Session
from openpyxl import load_workbook
file_pd=None
cp = {
    "account": st.secrets["snowflake2"]["account"],
    "user": st.secrets["snowflake2"]["user"],
    "password": st.secrets["snowflake2"]["password"],
    "role": st.secrets["snowflake2"]["role"],
    "warehouse": st.secrets["snowflake2"]["warehouse"],
    "database": st.secrets["snowflake2"]["database"],
    "schema": st.secrets["snowflake2"]["schema"]
}

def create_session():
    global curr_sess, curr_user, curr_role, curr_account, curr_env
    session = Session.builder.configs(cp).create()
    curr_sess = session
    curr_user=cp["user"].upper()
    curr_account=cp["account"].upper()
    curr_role=cp["role"].upper()
    session.query_tag="streamlit fileupload : "+curr_role
    return session

sess = create_session()

def read_sheet(obj_file,file_type='xlsx',str_sheetname="",col_list=None, date_col_list=False):
    dataframe=""    
    try:
        if obj_file is not None:
            if file_type == 'csv':
                dataframe = pd.read_csv(obj_file, usecols=col_list, parse_dates = date_col_list).applymap(lambda s: s.upper() if type(s) == str else s).fillna('')
            else:
                dataframe = pd.read_excel(obj_file,sheet_name = str_sheetname, usecols=col_list, parse_dates = date_col_list).applymap(lambda s: s.upper() if type(s) == str else s).fillna('')
        else:
            dataframe="Empty - Upload a file first"
            st.warning(dataframe)			
    except ValueError as e:
        st.error("Problem: "+ e.args[0])
    finally:
        return dataframe

def get_sheetnames(upload_file):
    wb = load_workbook(upload_file, read_only=True, keep_links=False)
    return wb.sheetnames

file_obj = st.file_uploader(
        "Choose File",
        help="Upload file in .xls .csv .xlsx format",
        type=['csv','xlsx','xls'])

if file_obj is not None:
    if file_obj.type == 'text/csv':
        file_pd=read_sheet(file_obj,'csv')
    else :
         excel_sheets = get_sheetnames(file_obj)
         excel_sheets_mod = ("Select Sheet",*excel_sheets)
         sheet = st.selectbox("Select Sheet",excel_sheets_mod, help = "Select a sheet from this list")
         if sheet != "Select Sheet":
            #st.write(sheet)
            file_pd = read_sheet(file_obj,'xls',sheet)
    if file_pd is not None:
        #st.write(file_pd)        
        tabname=st.text_input("Table  name: ")
        if tabname != "":
                df = sess.write_pandas(file_pd, tabname.upper(), auto_create_table=True)
                st.write(df.to_pandas())
                tbls = sess.sql("show tables like \'"+tabname.upper()+"%\'")
                tdf = pd.DataFrame(tbls.collect())
                st.write("Table in Snowflake")
                st.write(tdf)
    else:
        st.write('No data')

